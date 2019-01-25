import openpyxl 

wb=openpyxl.load_workbook('test.xlsx')
ws=wb.get_sheet_by_name('Sheet1')
ws['A1']='Tail row'
ws['B2']='Wide column'
ws.row_dimensions[1].height=70
ws.column_dimensions['C'].width=20
wb.save('test.xlsx')
