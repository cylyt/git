from openpyxl.styles import Font
import openpyxl
import os

wb=openpyxl.Workbook()
ws=wb.create_sheet('table')

for i in range (1,7):
    for j in range(1,7):
        if i==1:
            ws.cell(row=1,column=j+1).value=i*j
            ws.cell(row=1,column=j+1).font=Font(bold=True)
        if j==1:
            ws.cell(row=i+1,column=j).value=i*j
            ws.cell(row=i+1,column=j).font=Font(bold=True)
        ws.cell(row=i+1,column=j+1).value=i*j

wb.save('multiplication.xlsx')