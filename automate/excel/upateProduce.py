#coding=utf-8 
import openpyxl
from openpyxl.styles import Font




updateProducts={'芹菜':1.19,'大蒜':3.07,'柠檬':1.27}
wb=openpyxl.load_workbook('E:\\soft\\git\\automate\\12.6\\products.xlsx')
ws=wb.get_sheet_by_name('Sheet1')


ws['A1'].font=Font(bold=True)

for row in range(2,ws.max_row):
    produceName=ws.cell(row=row,column=1).value
    if  produceName in updateProducts.keys():
        updatePrice=updateProducts[produceName]
        ws.cell(row=row,column=2).value=updatePrice

wb.save('updateProduct.xlsx')